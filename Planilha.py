from tkinter import *
from openpyxl import *
from copy import *


wb = load_workbook(filename = 'Controle Receitas x Despesas.xlsx')
MA = wb.sheetnames[-1] #Da o nome do mês atual
MES = ['JANEIRO','FEVEREIRO','MARÇO','ABRIL','MAIO','JUNHO','JULHO','AGOSTO','SETEMBRO','OUTUBRO','NOVEMBRO','DEZEMBRO']
sheet = wb[MA]
ColunasPlanilha = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
LastRow = sheet.max_row

#Codigo copy cell pego de https://stackoverflow.com/questions/45347284/in-openpyxl-how-to-move-or-copy-a-cell-range-with-formatting-merged-cells-for  estudá-lo depois.

def copy_cell(source_cell, coord, tgt):
    tgt[coord].value = source_cell.value
    if source_cell.has_style:
        tgt[coord]._style = copy(source_cell._style)
    return tgt[coord]

def Mudar():
    if MES[11] == MA[:-3]:
        NovoAno = 'JANEIRO ' + str(int(MA[-2:]) + 1)
        wb.create_sheet(title = str(NovoAno))
        wb.save('Controle Receitas x Despesas.xlsx')
    else :
        Index = MES.index(MA[:-3])
        NovoMes = str(MES[int(Index + 1)]) + ' ' + str(int(MA[-2:]))
        wb.create_sheet(title = str(NovoMes))
        wb.save('Controle Receitas x Despesas.xlsx')
    MesNovo = wb.sheetnames[-1]
    sheetMesNovo = wb[MesNovo]
    sheetMesNovo.merge_cells('A1:H1')
    sheetMesNovo['A1'] = 'MÊS ' + str(MA[:-3]) + '/ ANO 20' + str(MA[-2:])
    for i in range(13):
        copy_cell(sheet[str(ColunasPlanilha[i]) + '2'],str(ColunasPlanilha[i]) + '2', sheetMesNovo)
    for i in range(13):
        copy_cell(sheet[str(ColunasPlanilha[i]) + '3'],str(ColunasPlanilha[i]) + '3', sheetMesNovo)
    copy_cell(sheetMesNovo['H' + str(LastRow - 1)], 'H3', sheet)
    sheetMesNovo.merge_cells('P6:Q6')
    sheetMesNovo['P6'] = 'LEGENDA'
    Mudarmes['state'] = 'disabled'
    MesCriadoLabel = Label(planilha, text='Novo mês criado')
    MesCriadoLabel.grid(row=500, column=50)
    wb.save('Controle Receitas x Despesas.xlsx')

def Salvar():
    if len(Data.get()) == 0 or len(Nome.get()) == 0 or len(Codigo.get()) == 0 or len(Valor.get()) == 0:
        FaltainfoLabel = Label(planilha, text='É necessário preencher todas as informações para salvar.')
        FaltainfoLabel.grid(row=550, column=50)
    else:
        LetrasdeSoma = ['F', 'G', 'J', 'K', 'L', 'F', 'M', 'N']
        for i in range(len(LetrasdeSoma)):
            sheet[str(LetrasdeSoma[i]) + str(LastRow)] = '=SUM(' + str(LetrasdeSoma[i]) + '3:' + str(LetrasdeSoma[i]) + str(LastRow) + ')'
        sheet.move_range('A' + str(LastRow) + ':' + 'AA' + str(LastRow), rows=1)
        for i in range(len(ColunasPlanilha)):
            copy_cell(sheet[str(ColunasPlanilha[i]) + str(LastRow-1)], str(ColunasPlanilha[i]) + str(LastRow) , sheet)
        sheet['J' + str(LastRow)] = '=IF(C' + str(LastRow) + '=1,F' + str(LastRow) + ',0)'
        for i in range(3):
            sheet[str(ColunasPlanilha[i + 10] + str(LastRow))] = '=IF(C' + str(LastRow) + '=' + str(i + 2) + ',G' + str(LastRow) + ',0)'
        GetData = Data.get()
        sheet['A' + str(LastRow)] = GetData
        GetNome = Nome.get()
        sheet['D' + str(LastRow)] = GetNome
        GetCodigo = Codigo.get()
        sheet['C' + str(LastRow)] = int(GetCodigo)
        if Codigo == 1:
            GetValor = Valor.get().replace(',', '.')
            sheet['F' + str(LastRow)] = '=' + GetValor
        else:
            GetValor = Valor.get().replace(',', '.')
            sheet['G' + str(LastRow)] = '=' + GetValor
        wb.save('Controle Receitas x Despesas.xlsx')


planilha = Tk()
planilha.title('Planilha')

# Colocar o título identificando o mês atual
TituloLabel = Label(planilha, text = 'O mês atual é ' + str(MA))
TituloLabel.grid(row = 50, column = 50)

#Colocar a data
dataLabel = Label(planilha, text = 'Data')
dataLabel.grid(row = 75, column = 50)
Data = Entry(planilha)
Data.grid(row = 75, column= 100)

#Colocar o nome da conta
NomeLabel = Label(planilha, text = 'Nome')
NomeLabel.grid(row = 100, column = 50)
Nome = Entry(planilha)
Nome.grid(row = 100, column= 100)

# Colocar o código da conta
CodigoLabel = Label(planilha, text = 'Código')
CodigoLabel.grid(row = 150, column = 50)
Codigo = Entry(planilha)
Codigo.grid(row = 150, column= 100)

# Colocar o valor da conta
ValorLabel = Label(planilha, text = 'Valor')
ValorLabel.grid(row = 200, column = 50)
Valor = Entry(planilha)
Valor.grid(row = 200, column= 100)

# Botão de salvar
Salvarlinha = Button(planilha, text = 'Salvar', width = 10, command=Salvar)
Salvarlinha.grid(row = 250, column = 50)

# Botão de mudar o mês
Mudarmes = Button(planilha, text = 'Mudar Mês', width = 10, command=Mudar)
Mudarmes.grid(row = 250, column = 100)

#Propriedades da janela
planilha.geometry('400x300+0+0')
planilha.mainloop()

'''
PARA FAZER:
- Fazer a função salvar funcionar devidamente (IMPORTANTE)
- Fazer com que ao mudar de mês apareça o novo mês no título
- Fazer com que seja possível ler um log das modificações desde que o aplicativo foi aberto
- Fazer com que mútiplos meses possam ser inseridos sem problema
- Fazer com que possa adicionar novas categorias de despesas ou receitas

'''